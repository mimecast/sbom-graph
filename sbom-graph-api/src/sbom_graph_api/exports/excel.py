"""Excel export utilities for graph data."""

from io import BytesIO
from typing import Any

from flask import Response as FlaskResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from sbom_graph_api.utils.validation import (
    sanitize_content_disposition,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def excel_response(
    buffer: BytesIO,
    filename: str,
) -> FlaskResponse:
    """Wrap a BytesIO Excel buffer in a downloadable Flask response."""
    safe_disp = sanitize_content_disposition(
        f"attachment; filename={filename}",
    )
    return FlaskResponse(
        buffer.getvalue(),
        mimetype=XLSX_MIME,
        headers={"Content-Disposition": safe_disp},
    )


def style_header_row(ws, num_cols: int) -> None:
    """Apply styling to the header row of a worksheet."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def auto_adjust_column_widths(ws) -> None:
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_source_impact_excel(
    impact: dict[str, Any],
    repo_url: str,
) -> BytesIO:
    """Create an Excel file for the source impact report."""
    packages = impact.get("packages", [])
    stats = impact.get("stats", {})

    wb = Workbook()
    ws = wb.active
    if ws is None:  # a freshly created Workbook always has an active sheet
        raise RuntimeError("Failed to create Excel worksheet")
    ws.title = "Source Impact"

    headers = [
        "Package",
        "Version",
        "Direct Dependants",
        "Transitive Dependants",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    for idx, pkg in enumerate(packages, start=2):
        ws.cell(row=idx, column=1, value=pkg.get("project_name", ""))
        ws.cell(row=idx, column=2, value=pkg.get("version", ""))
        ws.cell(row=idx, column=3, value=pkg.get("direct_dependants", 0))
        ws.cell(row=idx, column=4, value=pkg.get("transitive_dependants", 0))

    auto_adjust_column_widths(ws)

    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Source Repository")
    summary_ws.cell(row=1, column=2, value=repo_url)
    summary_ws.cell(row=2, column=1, value="Packages from Repo")
    summary_ws.cell(row=2, column=2, value=stats.get("packages_from_repo", 0))
    summary_ws.cell(row=3, column=1, value="Total Downstream Consumers")
    summary_ws.cell(row=3, column=2, value=stats.get("total_downstream_consumers", 0))
    summary_ws.cell(row=4, column=1, value="Affected Applications")
    summary_ws.cell(row=4, column=2, value=stats.get("affected_applications", 0))
    auto_adjust_column_widths(summary_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
