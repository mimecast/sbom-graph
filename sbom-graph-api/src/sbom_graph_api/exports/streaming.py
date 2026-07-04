"""Memory-safe streaming exporters (Phase 1).

Both writers stream page-by-page and never buffer the full result set in memory
(PERF-001 / SEC-003):

* :func:`stream_workbook_response` builds an ``openpyxl`` workbook in *write_only*
  mode (cells flushed to a 0600 temp file), then streams the file in chunks and
  unlinks it — on success and on error (SEC-006). A build error propagates before
  the HTTP response body commits, so the caller can surface a clean 500 (SEC-007).
* :func:`stream_json_response` emits the JSON document incrementally via
  ``stream_with_context``. A mid-stream source failure logs server-side and stops
  WITHOUT emitting the closing brace/stats, so the truncated document is detectably
  invalid and never forged-complete (SEC-007).
"""

from __future__ import annotations

import json
import logging
import os
import tempfile
from collections.abc import Iterable, Iterator
from dataclasses import dataclass, field
from typing import Any

from flask import Response, stream_with_context
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sbom_graph_api.utils.validation import sanitize_content_disposition

logger = logging.getLogger(__name__)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_CHUNK_BYTES = 65536

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="366092")
_HEADER_ALIGN = Alignment(horizontal="center")


@dataclass
class SheetSpec:
    """Specification for a single sheet in a multi-sheet workbook."""

    title: str
    headers: list[str]
    rows: Iterable[list[Any]]
    col_widths: list[float] | None = field(default=None)


def stream_workbook_response(
    headers: list[str],
    row_iter: Iterable[list[Any]],
    filename: str,
    sheet_title: str = "Report",
) -> Response:
    """Build a write-only xlsx from a row generator and stream it from a temp file.

    Memory stays flat: cells are flushed to disk as they are appended, and the
    finished file is streamed in fixed-size chunks. The temp file is always
    removed — after a successful stream or if the build raises.
    """
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)  # openpyxl writes by path; keep the 0600 file mkstemp created
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(sheet_title[:31] or "Report")
    try:
        worksheet.append(list(headers))
        for row in row_iter:
            worksheet.append(list(row))
        workbook.save(path)  # drives + closes the write-only row generator
    except BaseException:
        # row_iter raised before save(): the write-only worksheet's lazy
        # ``_write_rows`` generator is still suspended mid-``<sheetData>``. If left
        # to GC, its close writes the closing tag to an already-closed file and
        # raises "I/O operation on closed file" as an *unraisable* exception, which
        # pytest 9 then blames (flakily) on whatever test runs next. Closing the
        # worksheet here finalises that generator deterministically.
        try:
            if not worksheet.closed:
                worksheet.close()
        except Exception:  # pragma: no cover - best-effort resource release
            logger.debug("write_only worksheet close failed", exc_info=True)
        if os.path.exists(path):
            os.unlink(path)
        raise
    finally:
        try:
            workbook.close()
        except Exception:  # pragma: no cover - best-effort resource release
            logger.debug("write_only workbook close failed", exc_info=True)

    def generate() -> Iterator[bytes]:
        try:
            with open(path, "rb") as handle:
                while True:
                    chunk = handle.read(_CHUNK_BYTES)
                    if not chunk:
                        break
                    yield chunk
        finally:
            if os.path.exists(path):
                os.unlink(path)

    disposition = sanitize_content_disposition(f"attachment; filename={filename}").replace(
        "inline", "attachment"
    )
    return Response(
        generate(),
        mimetype=XLSX_MIME,
        headers={"Content-Disposition": disposition},
    )


def stream_multi_sheet_workbook_response(
    sheets: list[SheetSpec],
    filename: str,
) -> Response:
    """Build a write-only xlsx with multiple styled sheets and stream it from a temp file.

    Each SheetSpec supplies a title, headers (styled bold/white on #366092), a row
    iterable, and optional column widths. Temp-file lifecycle mirrors
    stream_workbook_response (created 0600, always unlinked).
    """
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    workbook = Workbook(write_only=True)
    worksheets = []
    try:
        for spec in sheets:
            ws = workbook.create_sheet(spec.title[:31] or "Sheet")
            worksheets.append((ws, spec))
            if spec.col_widths:
                for idx, width in enumerate(spec.col_widths, start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = width
            header_cells = []
            for h in spec.headers:
                cell = WriteOnlyCell(ws, value=h)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGN
                header_cells.append(cell)
            ws.append(header_cells)
            for row in spec.rows:
                ws.append(list(row))
        workbook.save(path)
    except BaseException:
        for ws, _ in worksheets:
            try:
                if not ws.closed:
                    ws.close()
            except Exception:  # pragma: no cover
                logger.debug("write_only worksheet close failed", exc_info=True)
        if os.path.exists(path):
            os.unlink(path)
        raise
    finally:
        try:
            workbook.close()
        except Exception:  # pragma: no cover
            logger.debug("write_only workbook close failed", exc_info=True)

    def generate() -> Iterator[bytes]:
        try:
            with open(path, "rb") as handle:
                while True:
                    chunk = handle.read(_CHUNK_BYTES)
                    if not chunk:
                        break
                    yield chunk
        finally:
            if os.path.exists(path):
                os.unlink(path)

    disposition = sanitize_content_disposition(f"attachment; filename={filename}").replace(
        "inline", "attachment"
    )
    return Response(
        generate(),
        mimetype=XLSX_MIME,
        headers={"Content-Disposition": disposition},
    )


def stream_json_response(
    meta: dict[str, Any],
    row_iter: Iterable[dict[str, Any]],
    filename: str,
    data_key: str = "data",
    stats: dict[str, Any] | None = None,
) -> Response:
    """Stream a JSON report document: ``{...meta, "data": [...], "stats": {...}}``.

    Rows are pulled lazily from ``row_iter`` so the full set is never materialised.
    On a mid-stream error the generator logs and stops without closing the document.
    """

    def generate() -> Iterator[bytes]:
        head = json.dumps(meta)
        if head.endswith("}"):
            head = head[:-1]
        sep = "," if meta else ""
        yield f'{head}{sep}"{data_key}":['.encode()
        first = True
        try:
            for row in row_iter:
                yield (("" if first else ",") + json.dumps(row, default=str)).encode()
                first = False
        except Exception:  # pragma: no cover - defensive
            # SEC-007: log server-side, stop WITHOUT closing -> truncated/invalid,
            # never a forged-complete document.
            logger.exception("report JSON stream failed; emitting truncated document")
            return
        if stats is not None:
            yield ('],"stats":' + json.dumps(stats) + "}").encode()
        else:
            yield b"]}"

    disposition = sanitize_content_disposition(f"attachment; filename={filename}").replace(
        "inline", "attachment"
    )
    return Response(
        stream_with_context(generate()),
        mimetype="application/json",
        headers={"Content-Disposition": disposition},
    )
