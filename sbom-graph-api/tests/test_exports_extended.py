"""Tests for source-impact Excel (the one survivor from the old create_* family)."""

from io import BytesIO

import openpyxl

from sbom_graph_api.exports.excel import create_source_impact_excel


def _load_wb(buffer: BytesIO) -> openpyxl.Workbook:
    buffer.seek(0)
    return openpyxl.load_workbook(buffer)


class TestCreateSourceImpactExcel:
    """Tests for create_source_impact_excel (still called from routes)."""

    def test_returns_bytesio(self):
        impact = {"packages": [], "stats": {}}
        result = create_source_impact_excel(impact, "https://github.com/example/repo")
        assert isinstance(result, BytesIO)

    def test_has_source_impact_and_summary_sheets(self):
        impact = {
            "packages": [
                {"project_name": "lib-a", "version": "1.0.0", "direct_dependants": 2, "transitive_dependants": 5},
            ],
            "stats": {"packages_from_repo": 1, "total_downstream_consumers": 5, "affected_applications": 3},
        }
        wb = _load_wb(create_source_impact_excel(impact, "https://github.com/x/y"))
        assert "Source Impact" in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_summary_contains_repo_url(self):
        impact = {"packages": [], "stats": {"packages_from_repo": 0}}
        repo_url = "https://github.com/example/repo"
        wb = _load_wb(create_source_impact_excel(impact, repo_url))
        summary = wb["Summary"]
        assert summary.cell(row=1, column=2).value == repo_url

    def test_data_rows_present(self):
        impact = {
            "packages": [
                {"project_name": "a", "version": "1.0", "direct_dependants": 1, "transitive_dependants": 2},
                {"project_name": "b", "version": "2.0", "direct_dependants": 3, "transitive_dependants": 4},
            ],
            "stats": {},
        }
        wb = _load_wb(create_source_impact_excel(impact, "https://example.com/repo"))
        ws = wb["Source Impact"]
        assert ws.max_row == 3  # header + 2 data rows
