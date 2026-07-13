"""Tests for export modules."""

from openpyxl import Workbook

from sbom_graph_api.exports.excel import (
    auto_adjust_column_widths,
    style_header_row,
)


class TestStyleHeaderRow:
    """Tests for style_header_row function."""

    def test_styles_header_cells(self):
        """Test header row styling is applied."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Header 1")
        ws.cell(row=1, column=2, value="Header 2")

        style_header_row(ws, 2)

        assert ws.cell(row=1, column=1).fill.start_color.rgb is not None


class TestAutoAdjustColumnWidths:
    """Tests for auto_adjust_column_widths function."""

    def test_adjusts_column_width(self):
        """Test column width adjustment."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Short")
        ws.cell(row=2, column=1, value="A much longer value here")

        auto_adjust_column_widths(ws)

        assert ws.column_dimensions["A"].width > 0
