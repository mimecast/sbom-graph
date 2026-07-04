"""Excel multi-sheet parity tests for streamed report exports (Phase 1.5 Step 4).

Each test drives a report's ``?format=excel`` path, loads the streamed xlsx via
openpyxl, and asserts: (1) the expected sheet names are present, (2) a sample
Summary value is correct, and (3) the main header row is styled (bold).
"""

from io import BytesIO
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook


def _load(response):
    """Drain a streamed Flask response and load it as an openpyxl workbook."""
    body = b"".join(response.response) if response.is_streamed else response.data
    return load_workbook(BytesIO(body))


def _header_is_styled(ws) -> bool:
    """True if the first header cell is bold (styled header row)."""
    return bool(ws.cell(row=1, column=1).font.bold)


def _summary_map(ws) -> dict:
    """Map a 2-column Metric/Value summary sheet into a dict (first col -> second)."""
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            out[row[0]] = row[1]
    return out


class TestMultiVersionDepsExcelParity:
    def test_excel_has_three_sheets_and_summary(self, client):
        mock = MagicMock()
        mock.get_library_version_usage.return_value = {
            "library": {"project_name": "lib", "total_versions": 2},
            "total_dependants": 3,
            "versions": [
                {
                    "version": "1.0",
                    "dependant_count": 2,
                    "is_internal": True,
                    "dependants": [
                        {
                            "project_name": "app",
                            "version": "9.0",
                            "project_group": "grp",
                            "is_internal": True,
                        }
                    ],
                },
                {"version": "2.0", "dependant_count": 0, "is_internal": False, "dependants": []},
            ],
        }
        with patch(
            "sbom_graph_api.routes.reports.dependencies.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(client.get("/reports/multi-version-deps/lib?format=excel"))
        assert wb.sheetnames == ["Version Usage", "Summary", "Version Summary"]
        assert _header_is_styled(wb["Version Usage"])
        summary = _summary_map(wb["Summary"])
        assert summary["Total Dependants"] == 3
        assert summary["Total Versions"] == 2


class TestMultiVersionSourcesExcelParity:
    def test_excel_has_three_sheets_and_summary(self, client):
        data = {
            "target": {"project_name": "proj", "version": "1.0", "scan_ids_count": 4},
            "multi_version_dependencies": [
                {
                    "dependency_project": "lib",
                    "version_count": 2,
                    "versions": [
                        {
                            "version": "1.0",
                            "contributing_applications": [{"project_name": "app", "version": "1.0"}],
                        },
                        {"version": "2.0", "contributing_applications": []},
                    ],
                }
            ],
        }
        mock = MagicMock()
        mock.find_multi_version_dependency_sources.return_value = data
        with patch(
            "sbom_graph_api.routes.reports.dependencies.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(client.get("/reports/multi-version-sources/proj/1.0?format=excel"))
        assert wb.sheetnames == [
            "Multi-Version Dependencies",
            "Summary",
            "Dependency Summary",
        ]
        assert _header_is_styled(wb["Multi-Version Dependencies"])
        summary = _summary_map(wb["Summary"])
        assert summary["Target Project"] == "proj"
        assert summary["Dependencies with Multiple Versions"] == 1
        assert summary["Total Conflicting Versions"] == 2


class TestNonSemverExcelParity:
    def test_excel_has_three_sheets_and_summary(self, client):
        mock = MagicMock()
        mock.find_non_semver_versions.return_value = [
            {"project_name": "p1", "version": "latest", "reason": "Non-standard", "labels": ["X"]},
            {"project_name": "p2", "version": "abc", "reason": "Git hash", "labels": []},
            {"project_name": "p3", "version": "snap", "reason": "Non-standard", "labels": []},
        ]
        with patch(
            "sbom_graph_api.routes.reports.dependencies.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(client.get("/reports/non-semver-versions?format=excel"))
        assert wb.sheetnames == ["Non-SemVer Versions", "Summary", "By Reason"]
        assert _header_is_styled(wb["Non-SemVer Versions"])
        summary = _summary_map(wb["Summary"])
        assert summary["Total Non-SemVer Versions"] == 3
        assert summary["Affected Projects"] == 3


class TestVersionDependenciesExcelParity:
    def test_excel_has_two_sheets_and_summary(self, client):
        mock = MagicMock()
        mock.is_project_semver_compliant.return_value = (True, [])
        mock.get_latest_semver_version.return_value = "2.0.0"
        mock.get_all_versions_of_project.return_value = ["1.0.0"]
        mock.get_transitive_dependencies_for_report.return_value = [
            {
                "depth": 1,
                "dependency_project": "lib-a",
                "dependency_version": "1.0.0",
                "is_internal": False,
            }
        ]
        with patch(
            "sbom_graph_api.routes.reports.dependencies.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(client.get("/reports/version-dependencies/proj/1.0.0?format=excel"))
        assert wb.sheetnames == ["Version Dependencies", "Summary"]
        assert _header_is_styled(wb["Version Dependencies"])
        summary = _summary_map(wb["Summary"])
        assert summary["Project Name"] == "proj"
        assert summary["Total Dependencies"] == 1
        assert summary["SemVer Compliant"] == "Yes"


class TestDependantsExcelParity:
    def test_excel_has_three_sheets_and_summary(self, client):
        mock = MagicMock()
        mock.find_version.return_value = {"id": "proj:1.0"}
        mock.get_dependants_with_partitions_and_paths.return_value = {
            "target": {"project_name": "proj", "version": "1.0"},
            "stats": {"total_dependants": 1, "max_partition": 2, "unique_projects": 1},
            "dependants": [
                {
                    "project_name": "dep",
                    "version": "3.0",
                    "partition": 2,
                    "max_path_edges": 2,
                    "path_count": 1,
                    "labels": ["INTERNAL"],
                    "paths": [["dep@3.0", "mid@1.0", "proj@1.0"]],
                }
            ],
        }
        with patch(
            "sbom_graph_api.routes.reports.dependencies.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(client.get("/reports/dependants/proj/1.0?format=excel"))
        assert wb.sheetnames == ["Dependants", "Dependency Paths", "Summary"]
        assert _header_is_styled(wb["Dependants"])
        # Dependency Paths sheet should contain the joined path.
        path_vals = [r[4] for r in wb["Dependency Paths"].iter_rows(min_row=2, values_only=True)]
        assert "dep@3.0 -> mid@1.0 -> proj@1.0" in path_vals
        summary = _summary_map(wb["Summary"])
        assert summary["Total Dependants"] == 1
        assert summary["Max Partition (Longest Path)"] == 2


class TestVulnerabilityDependantsExcelParity:
    def test_excel_has_three_sheets_and_summary(self, client):
        mock = MagicMock()
        mock.get_vulnerability_by_id.return_value = {
            "defect_id": "CVE-2021-1",
            "severity": "HIGH",
            "cvss_score": 7.5,
            "title": "Bad bug",
            "cwe_id": "CWE-79",
            "published_date": "2021-01-01",
            "description": "desc",
        }
        mock.get_vulnerability_dependants.return_value = [
            {
                "project_name": "dep",
                "version": "1.0",
                "partition": 1,
                "is_internal": True,
                "labels": ["INTERNAL"],
                "affected_by": [{"project_name": "lib", "version": "0.1"}],
            }
        ]
        with patch(
            "sbom_graph_api.routes.reports.vulnerabilities.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(client.get("/reports/vulnerability-dependants/CVE-2021-1?format=excel"))
        assert wb.sheetnames == ["Affected Dependants", "Vulnerability", "Summary"]
        assert _header_is_styled(wb["Affected Dependants"])
        vuln = _summary_map(wb["Vulnerability"])
        assert vuln["Severity"] == "HIGH"
        summary = _summary_map(wb["Summary"])
        assert summary["Total Dependants"] == 1
        assert summary["Unique Projects"] == 1


class TestEnrichmentCoverageExcelParity:
    def test_excel_has_summary_sheet(self, client):
        mock = MagicMock()
        mock.get_enrichment_coverage.return_value = {
            "total": 2,
            "recent": 1,
            "stale": 1,
            "never": 0,
            "recent_pct": 50.0,
            "stale_pct": 50.0,
            "never_pct": 0.0,
            "packages": [
                {
                    "purl": "pkg:x",
                    "project_name": "p",
                    "version_name": "1",
                    "last_enriched_at": "2026-01-01",
                    "status": "recent",
                }
            ],
        }
        with patch(
            "sbom_graph_api.routes.reports.vulnerabilities.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(client.get("/reports/enrichment-coverage?format=excel"))
        assert "Enrichment Coverage" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert _header_is_styled(wb["Enrichment Coverage"])
        summary = _summary_map(wb["Summary"])
        assert summary["Total Packages"] == 2


class TestVexCoverageExcelParity:
    def test_excel_has_summary_sheet(self, client):
        mock = MagicMock()
        mock.get_vulnerabilities_with_vex.return_value = [
            {
                "defect_id": "CVE-1",
                "severity": "LOW",
                "description": "d",
                "vex_status": "not_affected",
                "vex_count": 1,
            }
        ]
        mock.get_vex_coverage.return_value = {
            "total_vulnerabilities": 1,
            "with_vex": 1,
            "without_vex": 0,
            "coverage_percent": 100.0,
        }
        with patch(
            "sbom_graph_api.routes.reports.vulnerabilities.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(client.get("/reports/vex-coverage?format=excel"))
        assert "VEX Coverage" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert _header_is_styled(wb["VEX Coverage"])
        summary = _summary_map(wb["Summary"])
        assert summary["Total Vulnerabilities"] == 1


class TestLicenseDashboardExcelParity:
    def test_excel_has_summary_sheet(self, client):
        mock = MagicMock()
        mock.get_license_risk_stats.return_value = {
            "total_packages": 3,
            "categories": {"strong_copyleft": {"count": 1, "pct": 33.3}},
        }
        mock.get_license_risk_rows.side_effect = lambda internal_only=False, limit=None, offset=0, category=None: (
            [
                {
                    "category": "strong_copyleft",
                    "purl": "pkg:x",
                    "project_name": "p",
                    "version_name": "1",
                    "spdx_id": "GPL-3.0",
                    "license_name": "GPL",
                }
            ]
            if offset == 0
            else []
        )
        with patch(
            "sbom_graph_api.routes.reports.compliance.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(client.get("/reports/license-dashboard?format=excel"))
        assert wb.sheetnames == ["License Dashboard", "Summary"]
        assert _header_is_styled(wb["License Dashboard"])
        summary = _summary_map(wb["Summary"])
        assert summary["Total Packages"] == 3


class TestLicenseSummaryExcelParity:
    def test_excel_has_summary_sheet(self, client):
        mock = MagicMock()
        mock.get_license_summary.return_value = [
            {
                "project_group": "g",
                "project_name": "p",
                "version": "1",
                "purl": "pkg:x",
                "spdx_id": "MIT",
                "license_name": "MIT",
                "risk_category": "permissive",
            }
        ]
        with patch(
            "sbom_graph_api.routes.reports.compliance.get_falkordb_service",
            return_value=mock,
        ):
            wb = _load(
                client.get("/reports/license-summary?project_name=p&version_name=1&format=excel")
            )
        assert wb.sheetnames == ["License Summary", "Summary"]
        assert _header_is_styled(wb["License Summary"])
        summary = _summary_map(wb["Summary"])
        assert summary["Total Licenses"] == 1
        assert summary["Project"] == "p"
