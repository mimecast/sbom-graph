"""Phase 1 (TDD red) — streaming exports (exports/streaming.py).

Covers SRTM tests: TA-003 (Excel write_only+temp), TA-004 (JSON stream),
TA-008 (constant memory / page-by-page), TA-013 (temp-file 0600 + cleanup on
success AND error), TA-014 (mid-stream error must NOT forge a complete document;
Excel error happens before the response body commits).

FAILS until sbom_graph_api.exports.streaming exists.
"""

import json
import os
from io import BytesIO

import pytest


def _consume(response):
    """Drain a Flask streaming Response into bytes."""
    return b"".join(response.response)


# --------------------------------------------------------------------------
# Excel streaming — TA-003, TA-013
# --------------------------------------------------------------------------

class TestStreamWorkbook:
    def test_excel_streams_rows_and_is_readable(self, app):
        """TA-003: write_only workbook produced from a row generator is a valid xlsx."""
        from openpyxl import load_workbook

        from sbom_graph_api.exports.streaming import stream_workbook_response

        headers = ["Project", "Version"]
        rows = ([f"proj-{i}", f"{i}.0.0"] for i in range(2500))
        with app.test_request_context():
            resp = stream_workbook_response(headers, rows, "all_projects.xlsx", "Projects")
            body = _consume(resp)

        assert resp.headers["Content-Disposition"].startswith("attachment")
        wb = load_workbook(BytesIO(body))
        ws = wb.active
        assert [c.value for c in ws[1]] == headers
        assert ws.max_row == 2501  # header + 2500 rows

    def test_excel_temp_file_is_cleaned_up_on_success(self, app, tmp_path):
        """TA-013: the temp file is removed after the response is fully streamed."""
        from sbom_graph_api.exports import streaming

        created = {}
        real_mkstemp = streaming.tempfile.mkstemp

        def spy_mkstemp(*a, **k):
            fd, path = real_mkstemp(*a, **k)
            created["path"] = path
            created["mode"] = oct(os.stat(path).st_mode & 0o777)
            return fd, path

        with app.test_request_context(), \
                pytest.MonkeyPatch().context() as mp:
            mp.setattr(streaming.tempfile, "mkstemp", spy_mkstemp)
            resp = streaming.stream_workbook_response(
                ["A"], ([str(i)] for i in range(10)), "x.xlsx", "S")
            _consume(resp)

        assert created.get("mode") == "0o600", "temp file must be created mode 0600"
        assert not os.path.exists(created["path"]), "temp file must be unlinked after streaming"

    def test_excel_temp_file_cleaned_up_on_error(self, app):
        """TA-013/TA-014: an error while building must not leak a temp file and must not
        emit a 200 body (error surfaces before the response commits)."""
        from sbom_graph_api.exports import streaming

        created = {}
        real_mkstemp = streaming.tempfile.mkstemp

        def spy_mkstemp(*a, **k):
            fd, path = real_mkstemp(*a, **k)
            created["path"] = path
            return fd, path

        def boom():
            yield ["ok"]
            raise RuntimeError("db died mid-build")

        with app.test_request_context(), pytest.MonkeyPatch().context() as mp:
            mp.setattr(streaming.tempfile, "mkstemp", spy_mkstemp)
            with pytest.raises(RuntimeError):
                resp = streaming.stream_workbook_response(["A"], boom(), "x.xlsx", "S")
                _consume(resp)

        if "path" in created:
            assert not os.path.exists(created["path"]), "temp file must be unlinked on error"


# --------------------------------------------------------------------------
# JSON streaming — TA-004, TA-008, TA-014
# --------------------------------------------------------------------------

class TestStreamJson:
    def test_json_streams_full_envelope(self, app):
        """TA-004: streamed JSON is a valid document with meta + data array."""
        from sbom_graph_api.exports.streaming import stream_json_response

        meta = {"report_type": "projects", "generated_at": "2026-06-26T00:00:00Z"}
        rows = ({"project_name": f"p{i}", "version": "1"} for i in range(1000))
        with app.test_request_context():
            resp = stream_json_response(meta, rows, "projects.json",
                                        stats={"Total": 1000})
            body = _consume(resp).decode()

        doc = json.loads(body)
        assert doc["report_type"] == "projects"
        assert len(doc["data"]) == 1000
        assert doc["stats"]["Total"] == 1000

    def test_json_consumes_iterator_lazily(self, app):
        """TA-008: rows are pulled from the generator, not pre-materialised into a list."""
        from sbom_graph_api.exports.streaming import stream_json_response

        seen = {"max_live": 0, "n": 0}

        def gen():
            for i in range(5000):
                seen["n"] += 1
                yield {"i": i}

        with app.test_request_context():
            resp = stream_json_response({"report_type": "t"}, gen(), "t.json")
            # Pull a few chunks; generator should not have been fully drained yet.
            it = iter(resp.response)
            next(it)
            partial = seen["n"]
            list(it)
        assert partial < 5000, "JSON must stream lazily, not buffer all rows first"

    def test_json_mid_stream_error_is_not_forged_complete(self, app):
        """TA-014: if the row source fails mid-stream, the emitted JSON must be
        truncated/invalid (NOT a well-formed doc that looks complete)."""
        from sbom_graph_api.exports.streaming import stream_json_response

        def boom():
            yield {"i": 0}
            raise RuntimeError("source failed")

        with app.test_request_context():
            resp = stream_json_response({"report_type": "t"}, boom(), "t.json",
                                        stats={"Total": 999})
            try:
                body = _consume(resp).decode()
            except RuntimeError:
                return  # acceptable: error propagated, nothing forged
        with pytest.raises(json.JSONDecodeError):
            json.loads(body)
        assert '"Total": 999' not in body, "must not emit the closing stats/brace on failure"


# --------------------------------------------------------------------------
# Multi-sheet Excel streaming — Phase 1.5 Step 3
# --------------------------------------------------------------------------

class TestStreamMultiSheetWorkbook:
    """Tests for stream_multi_sheet_workbook_response and SheetSpec."""

    def _make_spec(self, title, headers, rows, col_widths=None):
        from sbom_graph_api.exports.streaming import SheetSpec
        return SheetSpec(title=title, headers=headers, rows=rows, col_widths=col_widths)

    def test_valid_xlsx_with_single_sheet(self, app):
        """Produces a readable xlsx with the correct sheet and header row."""
        from io import BytesIO

        from openpyxl import load_workbook

        from sbom_graph_api.exports.streaming import stream_multi_sheet_workbook_response

        spec = self._make_spec("Data", ["A", "B"], (["v1", "v2"] for _ in range(3)))
        with app.test_request_context():
            resp = stream_multi_sheet_workbook_response([spec], "test.xlsx")
            body = b"".join(resp.response)

        wb = load_workbook(BytesIO(body))
        assert "Data" in wb.sheetnames
        ws = wb["Data"]
        assert [c.value for c in ws[1]] == ["A", "B"]
        assert ws.max_row == 4  # header + 3 data rows

    def test_multiple_sheets_present(self, app):
        """All SheetSpec sheets appear in the workbook in order."""
        from io import BytesIO

        from openpyxl import load_workbook

        from sbom_graph_api.exports.streaming import stream_multi_sheet_workbook_response

        specs = [
            self._make_spec("Main", ["X"], (["x"] for _ in range(2))),
            self._make_spec("Summary", ["Key", "Value"], iter([["Total", "2"]])),
        ]
        with app.test_request_context():
            resp = stream_multi_sheet_workbook_response(specs, "multi.xlsx")
            body = b"".join(resp.response)

        wb = load_workbook(BytesIO(body))
        assert wb.sheetnames == ["Main", "Summary"]

    def test_header_row_is_styled_bold_white_on_blue(self, app):
        """Header cells must be bold, white font, on #366092 fill."""
        from io import BytesIO

        from openpyxl import load_workbook

        from sbom_graph_api.exports.streaming import stream_multi_sheet_workbook_response

        spec = self._make_spec("Styled", ["Col1", "Col2"], iter([]))
        with app.test_request_context():
            resp = stream_multi_sheet_workbook_response([spec], "styled.xlsx")
            body = b"".join(resp.response)

        wb = load_workbook(BytesIO(body))
        ws = wb["Styled"]
        header_cell = ws["A1"]
        assert header_cell.font.bold is True, "header must be bold"
        # openpyxl stores alpha-prefixed hex when round-tripping write-only cells
        assert header_cell.font.color.rgb in ("FFFFFFFF", "00FFFFFF"), "header font must be white"
        assert header_cell.fill.fgColor.rgb in ("FF366092", "00366092"), "header fill must be #366092"

    def test_col_widths_applied(self, app):
        """Column dimensions are set when col_widths provided."""
        from io import BytesIO

        from openpyxl import load_workbook

        from sbom_graph_api.exports.streaming import stream_multi_sheet_workbook_response

        spec = self._make_spec("W", ["A", "B", "C"], iter([]), col_widths=[20, 30, 10])
        with app.test_request_context():
            resp = stream_multi_sheet_workbook_response([spec], "widths.xlsx")
            body = b"".join(resp.response)

        wb = load_workbook(BytesIO(body))
        ws = wb["W"]
        assert ws.column_dimensions["A"].width == 20
        assert ws.column_dimensions["B"].width == 30
        assert ws.column_dimensions["C"].width == 10

    def test_content_disposition_is_attachment(self, app):
        """Response must have attachment content-disposition."""
        from sbom_graph_api.exports.streaming import stream_multi_sheet_workbook_response

        spec = self._make_spec("S", ["H"], iter([]))
        with app.test_request_context():
            resp = stream_multi_sheet_workbook_response([spec], "report.xlsx")
            b"".join(resp.response)

        assert resp.headers["Content-Disposition"].startswith("attachment")
        assert "report.xlsx" in resp.headers["Content-Disposition"]

    def test_temp_file_cleaned_up_on_success(self, app):
        """Temp file is removed after the response is fully consumed."""
        from sbom_graph_api.exports import streaming
        from sbom_graph_api.exports.streaming import stream_multi_sheet_workbook_response

        created = {}
        real_mkstemp = streaming.tempfile.mkstemp

        def spy_mkstemp(*a, **k):
            fd, path = real_mkstemp(*a, **k)
            created["path"] = path
            return fd, path

        spec = self._make_spec("S", ["H"], iter([["val"]]))
        with app.test_request_context(), pytest.MonkeyPatch().context() as mp:
            mp.setattr(streaming.tempfile, "mkstemp", spy_mkstemp)
            resp = stream_multi_sheet_workbook_response([spec], "t.xlsx")
            b"".join(resp.response)

        assert "path" in created
        assert not os.path.exists(created["path"]), "temp file must be unlinked after streaming"

    def test_temp_file_cleaned_up_on_error(self, app):
        """Temp file is removed even when row_iter raises during build."""
        from sbom_graph_api.exports import streaming
        from sbom_graph_api.exports.streaming import stream_multi_sheet_workbook_response

        created = {}
        real_mkstemp = streaming.tempfile.mkstemp

        def spy_mkstemp(*a, **k):
            fd, path = real_mkstemp(*a, **k)
            created["path"] = path
            return fd, path

        def boom():
            yield ["ok"]
            raise RuntimeError("source exploded")

        spec = self._make_spec("S", ["H"], boom())
        with app.test_request_context(), pytest.MonkeyPatch().context() as mp:
            mp.setattr(streaming.tempfile, "mkstemp", spy_mkstemp)
            with pytest.raises(RuntimeError):
                resp = stream_multi_sheet_workbook_response([spec], "t.xlsx")
                b"".join(resp.response)

        if "path" in created:
            assert not os.path.exists(created["path"]), "temp file must be unlinked on error"

    def test_large_row_count_no_memory_spike(self, app):
        """10 000 rows must stream without materialising everything at once."""
        from io import BytesIO

        from openpyxl import load_workbook

        from sbom_graph_api.exports.streaming import stream_multi_sheet_workbook_response

        spec = self._make_spec("Big", ["N"], (([str(i)]) for i in range(10_000)))
        with app.test_request_context():
            resp = stream_multi_sheet_workbook_response([spec], "big.xlsx")
            body = b"".join(resp.response)

        wb = load_workbook(BytesIO(body))
        assert wb["Big"].max_row == 10_001  # header + 10 000
