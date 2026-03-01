"""Extended tests for Excel export functions."""

from io import BytesIO

import openpyxl
import pytest

from appsec_data_views.exports.excel import (
    create_applications_excel,
    create_centrality_excel,
    create_dependants_report_excel,
    create_multi_version_dependency_report_excel,
    create_multi_version_deps_excel,
    create_non_semver_report_excel,
    create_vulnerabilities_excel,
    create_vulnerability_dependants_excel,
    create_version_dependencies_report_excel,
)


def _load_wb(buffer: BytesIO) -> openpyxl.Workbook:
    buffer.seek(0)
    return openpyxl.load_workbook(buffer)


class TestCreateApplicationsExcel:
    """Tests for create_applications_excel."""

    def test_returns_valid_excel(self):
        from unittest.mock import MagicMock
        mock_service = MagicMock()
        mock_service.get_all_applications.return_value = [
            {"project_name": "app-a", "version": "1.0.0", "scan_id": "s1",
             "app_id": "a1", "public_id": "p1", "repo_url": "https://git", "is_internal": True},
        ]
        buffer = create_applications_excel(mock_service, limit=100)
        wb = _load_wb(buffer)
        assert "Applications" in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_internal_only_title(self):
        from unittest.mock import MagicMock
        mock_service = MagicMock()
        mock_service.get_all_applications.return_value = []
        buffer = create_applications_excel(mock_service, internal_only=True)
        wb = _load_wb(buffer)
        assert any("Internal" in name for name in wb.sheetnames)

    def test_latest_only_title(self):
        from unittest.mock import MagicMock
        mock_service = MagicMock()
        mock_service.get_all_applications.return_value = []
        buffer = create_applications_excel(mock_service, latest_only=True)
        wb = _load_wb(buffer)
        assert any("Latest" in name for name in wb.sheetnames)


class TestCreateVersionDependenciesReportExcel:
    """Tests for create_version_dependencies_report_excel."""

    def test_with_dependencies(self):
        deps = [
            {"depth": 1, "dependency_project": "lib-a", "dependency_version": "1.0",
             "is_internal": True},
            {"depth": 2, "dependency_project": "lib-b", "dependency_version": "2.0",
             "is_internal": False},
        ]
        buffer = create_version_dependencies_report_excel(
            "my-proj", "1.0.0", deps, True, "1.0.0",
        )
        wb = _load_wb(buffer)
        ws = wb["Version Dependencies"]
        assert ws.cell(row=1, column=1).value == "Depth"
        assert "Summary" in wb.sheetnames

    def test_empty_dependencies(self):
        buffer = create_version_dependencies_report_excel(
            "my-proj", "1.0.0", [], False, None,
        )
        wb = _load_wb(buffer)
        ws = wb["Version Dependencies"]
        assert ws.cell(row=2, column=2).value == "(no dependencies)"

    def test_summary_has_semver_info(self):
        deps = [{"depth": 1, "dependency_project": "lib", "dependency_version": "1.0", "is_internal": False}]
        buffer = create_version_dependencies_report_excel(
            "proj", "1.0.0", deps, True, "2.0.0",
        )
        wb = _load_wb(buffer)
        summary = wb["Summary"]
        values = [summary.cell(row=r, column=2).value for r in range(1, 15)]
        assert "2.0.0" in values


class TestCreateMultiVersionDependencyReportExcel:
    """Tests for multi-version dependency source report."""

    def test_with_data(self):
        data = {
            "target": {"project_name": "proj", "version": "1.0", "scan_ids_count": 2},
            "multi_version_dependencies": [
                {
                    "dependency_project": "lib",
                    "version_count": 2,
                    "versions": [
                        {"version": "1.0", "scan_ids_intersection": ["s1"],
                         "contributing_applications": [{"project_name": "app-a", "version": "1.0"}]},
                        {"version": "2.0", "scan_ids_intersection": [],
                         "contributing_applications": []},
                    ],
                },
            ],
        }
        buffer = create_multi_version_dependency_report_excel(data)
        wb = _load_wb(buffer)
        assert "Multi-Version Dependencies" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert "Dependency Summary" in wb.sheetnames

    def test_empty_data(self):
        data = {"target": {}, "multi_version_dependencies": []}
        buffer = create_multi_version_dependency_report_excel(data)
        wb = _load_wb(buffer)
        assert wb is not None


class TestCreateMultiVersionDepsExcel:
    """Tests for library version usage Excel."""

    def test_with_dependants(self):
        data = {
            "library": {"project_name": "my-lib", "total_versions": 2},
            "total_dependants": 3,
            "versions": [
                {"version": "1.0", "dependant_count": 2, "is_internal": True,
                 "dependants": [
                     {"project_name": "app-a", "version": "1.0", "project_group": "g", "is_internal": True},
                     {"project_name": "app-b", "version": "2.0", "project_group": "g", "is_internal": False},
                 ]},
                {"version": "2.0", "dependant_count": 0, "is_internal": False, "dependants": []},
            ],
        }
        buffer = create_multi_version_deps_excel(data)
        wb = _load_wb(buffer)
        assert "Version Usage" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert "Version Summary" in wb.sheetnames


class TestCreateNonSemverReportExcel:
    """Tests for non-SemVer report Excel."""

    def test_with_data(self):
        data = [
            {"project_name": "proj-a", "version": "latest", "reason": "No numeric component",
             "labels": ["Version"]},
            {"project_name": "proj-b", "version": "20230101", "reason": "Date-based version (YYYYMMDD)",
             "labels": ["Version", "INTERNAL"]},
        ]
        buffer = create_non_semver_report_excel(data)
        wb = _load_wb(buffer)
        assert "Non-SemVer Versions" in wb.sheetnames
        assert "By Reason" in wb.sheetnames

    def test_empty_data(self):
        buffer = create_non_semver_report_excel([])
        wb = _load_wb(buffer)
        summary = wb["Summary"]
        assert summary.cell(row=1, column=2).value == 0


class TestCreateDependantsReportExcel:
    """Tests for dependants report Excel."""

    def test_with_dependants_and_paths(self):
        data = {
            "target": {"project_name": "lib", "version": "1.0"},
            "stats": {"total_dependants": 1, "max_partition": 2, "unique_projects": 1},
            "dependants": [
                {"project_name": "app-a", "version": "1.0", "partition": 1,
                 "max_path_edges": 1, "path_count": 1,
                 "labels": ["Version", "INTERNAL"],
                 "paths": [["app-a@1.0", "lib@1.0"]]},
            ],
        }
        buffer = create_dependants_report_excel(data, longest_only=True)
        wb = _load_wb(buffer)
        assert "Dependants" in wb.sheetnames
        assert "Dependency Paths" in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_empty_dependants(self):
        data = {
            "target": {"project_name": "lib", "version": "1.0"},
            "stats": {"total_dependants": 0, "max_partition": 0, "unique_projects": 0},
            "dependants": [],
        }
        buffer = create_dependants_report_excel(data)
        wb = _load_wb(buffer)
        assert wb is not None


class TestCreateVulnerabilitiesExcel:
    """Tests for vulnerabilities Excel."""

    def test_with_vulnerabilities(self):
        vulns = [
            {"defect_id": "CVE-2024-001", "severity": "CRITICAL", "cvss_score": 9.8,
             "title": "Critical Bug", "cwe_id": "CWE-79", "published_date": "2024-01-01",
             "affected_versions": [{"project_name": "lib", "version": "1.0"}]},
            {"defect_id": "CVE-2024-002", "severity": "LOW", "cvss_score": 2.0,
             "title": "Info Leak", "cwe_id": None, "published_date": None,
             "affected_versions": []},
        ]
        buffer = create_vulnerabilities_excel(vulns, internal_only=False)
        wb = _load_wb(buffer)
        assert "Vulnerabilities" in wb.sheetnames
        summary = wb["Summary"]
        assert summary.cell(row=1, column=2).value == 2

    def test_empty_vulnerabilities(self):
        buffer = create_vulnerabilities_excel([])
        wb = _load_wb(buffer)
        assert wb is not None


class TestCreateVulnerabilityDependantsExcel:
    """Tests for vulnerability dependants Excel."""

    def test_with_data(self):
        vuln = {"defect_id": "CVE-2024-001", "severity": "HIGH", "cvss_score": 7.5,
                "title": "XSS", "cwe_id": "CWE-79", "published_date": "2024-01-01",
                "description": "Cross-site scripting vulnerability"}
        dependants = [
            {"project_name": "app-a", "version": "1.0", "partition": 1,
             "is_internal": True, "affected_by": [{"project_name": "lib", "version": "1.0"}]},
            {"project_name": "app-b", "version": "2.0", "partition": 2,
             "is_internal": False, "affected_by": [{"project_name": "lib", "version": "1.0"}]},
        ]
        buffer = create_vulnerability_dependants_excel(vuln, dependants)
        wb = _load_wb(buffer)
        assert "Affected Dependants" in wb.sheetnames
        assert "Vulnerability" in wb.sheetnames
        assert "Summary" in wb.sheetnames


class TestCreateCentralityExcel:
    """Tests for centrality Excel."""

    def test_with_data(self):
        data = [
            {"inDegree": 10, "outDegree": 5, "project_name": "lib-a",
             "project_group": "com.example", "version_name": "1.0.0"},
            {"inDegree": 8, "outDegree": 3, "project_name": "lib-b",
             "project_group": "com.example", "version_name": "2.0.0"},
        ]
        buffer = create_centrality_excel(data)
        wb = _load_wb(buffer)
        assert "Centrality" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        summary = wb["Summary"]
        assert summary.cell(row=1, column=2).value == 2

    def test_empty_data(self):
        buffer = create_centrality_excel([])
        wb = _load_wb(buffer)
        assert wb is not None
