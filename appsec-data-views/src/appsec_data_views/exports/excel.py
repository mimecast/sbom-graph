"""Excel export utilities for graph data.

This module provides functions to generate Excel files from graph data
using openpyxl and pandas.
"""

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from appsec_data_views.services.falkordb_service import FalkorDBService, get_falkordb_service


def style_header_row(ws, num_cols: int) -> None:
    """Apply styling to the header row of a worksheet."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def auto_adjust_column_widths(ws) -> None:
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_version_dependencies_excel(
    project_name: str,
    service: FalkorDBService | None = None,
    internal_only: bool = False,
) -> BytesIO:
    """Create an Excel file with version to dependant project versions.

    Args:
        project_name: The project name to export
        service: FalkorDB service instance
        internal_only: If True, only include internal-labeled nodes

    Returns:
        BytesIO buffer containing the Excel file
    """
    if service is None:
        service = get_falkordb_service()

    # Get all versions of the project
    versions = service.get_all_versions_of_project(project_name, internal_only)

    # Get all direct dependants
    dependants = service.get_direct_dependants(project_name, internal_only=internal_only)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Version Dependencies"

    # Create DataFrame from dependants data
    if dependants:
        df = pd.DataFrame(dependants)
        df = df.rename(
            columns={
                "target_version": "Version",
                "dependant_project": "Dependant Project",
                "dependant_version": "Dependant Version",
            }
        )
        df = df[["Version", "Dependant Project", "Dependant Version"]]
        df = df.sort_values(["Version", "Dependant Project", "Dependant Version"])

        # Write to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        style_header_row(ws, len(df.columns))
    else:
        # No dependants found
        ws.cell(row=1, column=1, value="Version")
        ws.cell(row=1, column=2, value="Dependant Project")
        ws.cell(row=1, column=3, value="Dependant Version")
        style_header_row(ws, 3)

        row = 2
        for version in versions:
            ws.cell(row=row, column=1, value=version)
            ws.cell(row=row, column=2, value="(no dependants)")
            ws.cell(row=row, column=3, value="-")
            row += 1

    auto_adjust_column_widths(ws)

    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Project Name")
    summary_ws.cell(row=1, column=2, value=project_name)
    summary_ws.cell(row=2, column=1, value="Total Versions")
    summary_ws.cell(row=2, column=2, value=len(versions))
    summary_ws.cell(row=3, column=1, value="Total Dependant Relationships")
    summary_ws.cell(row=3, column=2, value=len(dependants))

    # Count unique dependants
    unique_dependants = len({(d["dependant_project"], d["dependant_version"]) for d in dependants})
    summary_ws.cell(row=4, column=1, value="Unique Dependant Versions")
    summary_ws.cell(row=4, column=2, value=unique_dependants)

    auto_adjust_column_widths(summary_ws)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_version_dependencies_report_excel(
    project_name: str,
    version_name: str,
    dependencies: list[dict],
    is_semver_compliant: bool,
    latest_version: str | None,
    internal_only: bool = False,
    max_depth: int = 10,
) -> BytesIO:
    """Create an Excel file for the version dependencies report.

    Args:
        project_name: The project name
        version_name: The version being analyzed
        dependencies: List of dependency records with depth, dependency_project, dependency_version, is_internal
        is_semver_compliant: Whether the project is fully SemVer compliant
        latest_version: The latest SemVer version if compliant
        internal_only: If True, data was filtered to internal-labeled nodes only
        max_depth: Maximum depth that was used for traversal

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Version Dependencies"

    # Create DataFrame from dependencies data
    if dependencies:
        df = pd.DataFrame(dependencies)
        df = df.rename(
            columns={
                "depth": "Depth",
                "dependency_project": "Dependency Project",
                "dependency_version": "Dependency Version",
                "is_internal": "Is INTERNAL",
            }
        )
        # Ensure columns exist and order
        cols = ["Depth", "Dependency Project", "Dependency Version", "Is Internal"]
        for col in cols:
            if col not in df.columns:
                df[col] = ""
        df = df[cols]
        df = df.sort_values(["Depth", "Dependency Project", "Dependency Version"])

        # Write to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        style_header_row(ws, len(df.columns))
    else:
        # No dependencies found
        ws.cell(row=1, column=1, value="Depth")
        ws.cell(row=1, column=2, value="Dependency Project")
        ws.cell(row=1, column=3, value="Dependency Version")
        ws.cell(row=1, column=4, value="Is Internal")
        style_header_row(ws, 4)

        ws.cell(row=2, column=1, value="-")
        ws.cell(row=2, column=2, value="(no dependencies)")
        ws.cell(row=2, column=3, value="-")
        ws.cell(row=2, column=4, value="-")

    auto_adjust_column_widths(ws)

    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    row = 1

    summary_ws.cell(row=row, column=1, value="Project Name")
    summary_ws.cell(row=row, column=2, value=project_name)
    row += 1

    summary_ws.cell(row=row, column=1, value="Version")
    summary_ws.cell(row=row, column=2, value=version_name)
    row += 1

    summary_ws.cell(row=row, column=1, value="Filter Mode")
    summary_ws.cell(row=row, column=2, value="Internal Only" if internal_only else "All")
    row += 1

    summary_ws.cell(row=row, column=1, value="Max Depth Setting")
    summary_ws.cell(row=row, column=2, value=max_depth)
    row += 1

    summary_ws.cell(row=row, column=1, value="Total Dependencies")
    summary_ws.cell(row=row, column=2, value=len(dependencies))
    row += 1

    # Count unique dependencies
    unique_deps = (
        len({(d["dependency_project"], d["dependency_version"]) for d in dependencies})
        if dependencies
        else 0
    )
    summary_ws.cell(row=row, column=1, value="Unique Dependencies")
    summary_ws.cell(row=row, column=2, value=unique_deps)
    row += 1

    # Direct dependencies (depth 1)
    direct_deps = sum(1 for d in dependencies if d.get("depth") == 1) if dependencies else 0
    summary_ws.cell(row=row, column=1, value="Direct Dependencies")
    summary_ws.cell(row=row, column=2, value=direct_deps)
    row += 1

    # Max depth reached
    max_depth_reached = max(d.get("depth", 0) for d in dependencies) if dependencies else 0
    summary_ws.cell(row=row, column=1, value="Max Depth Reached")
    summary_ws.cell(row=row, column=2, value=max_depth_reached)
    row += 1

    # SemVer compliance info
    row += 1
    summary_ws.cell(row=row, column=1, value="SemVer Compliant")
    summary_ws.cell(row=row, column=2, value="Yes" if is_semver_compliant else "No")
    row += 1

    if is_semver_compliant and latest_version:
        summary_ws.cell(row=row, column=1, value="Latest SemVer Version")
        summary_ws.cell(row=row, column=2, value=latest_version)
        row += 1

    auto_adjust_column_widths(summary_ws)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_all_projects_excel(
    service: FalkorDBService | None = None,
    limit: int = 10000,
    internal_only: bool = False,
) -> BytesIO:
    """Create an Excel file with all projects and their versions.

    Args:
        service: FalkorDB service instance
        limit: Maximum number of results
        internal_only: If True, only include internal-labeled nodes

    Returns:
        BytesIO buffer containing the Excel file
    """
    if service is None:
        service = get_falkordb_service()

    projects = service.get_all_projects(limit, internal_only)

    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Projects" if internal_only else "All Projects"

    # Headers
    ws.cell(row=1, column=1, value="Project Name")
    ws.cell(row=1, column=2, value="Version")
    style_header_row(ws, 2)

    # Data
    for idx, project in enumerate(projects, start=2):
        ws.cell(row=idx, column=1, value=project["project_name"])
        ws.cell(row=idx, column=2, value=project["version"])

    auto_adjust_column_widths(ws)

    # Add summary
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Total Project Versions")
    summary_ws.cell(row=1, column=2, value=len(projects))

    # Count unique projects
    unique_projects = len({p["project_name"] for p in projects})
    summary_ws.cell(row=2, column=1, value="Unique Projects")
    summary_ws.cell(row=2, column=2, value=unique_projects)

    summary_ws.cell(row=3, column=1, value="Filter")
    summary_ws.cell(row=3, column=2, value="Internal Only" if internal_only else "All")

    auto_adjust_column_widths(summary_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_applications_excel(
    service: FalkorDBService | None = None,
    limit: int = 10000,
    internal_only: bool = False,
    latest_only: bool = False,
) -> BytesIO:
    """Create an Excel file with all applications and their versions.

    Args:
        service: FalkorDB service instance
        limit: Maximum number of results
        internal_only: If True, only include internal-labeled nodes
        latest_only: If True, only include the latest version per application

    Returns:
        BytesIO buffer containing the Excel file
    """
    if service is None:
        service = get_falkordb_service()

    applications = service.get_all_applications(limit, internal_only, latest_only)

    wb = Workbook()
    ws = wb.active

    title_parts = []
    if internal_only:
        title_parts.append("Internal")
    if latest_only:
        title_parts.append("Latest")
    title_parts.append("Applications")
    ws.title = " ".join(title_parts)

    # Headers
    headers = [
        "Project Name",
        "Version",
        "Scan ID",
        "App ID",
        "Public ID",
        "Repo URL",
        "Is Internal",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    # Data
    for idx, app in enumerate(applications, start=2):
        ws.cell(row=idx, column=1, value=app["project_name"])
        ws.cell(row=idx, column=2, value=app["version"])
        ws.cell(row=idx, column=3, value=app.get("scan_id"))
        ws.cell(row=idx, column=4, value=app.get("app_id"))
        ws.cell(row=idx, column=5, value=app.get("public_id"))
        ws.cell(row=idx, column=6, value=app.get("repo_url"))
        ws.cell(row=idx, column=7, value="Yes" if app.get("is_internal") else "No")

    auto_adjust_column_widths(ws)

    # Add summary
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Total Applications")
    summary_ws.cell(row=1, column=2, value=len(applications))

    # Count unique projects
    unique_projects = len({a["project_name"] for a in applications})
    summary_ws.cell(row=2, column=1, value="Unique Applications")
    summary_ws.cell(row=2, column=2, value=unique_projects)

    summary_ws.cell(row=3, column=1, value="Filter Mode")
    filter_mode = "Internal Only" if internal_only else "All"
    summary_ws.cell(row=3, column=2, value=filter_mode)

    summary_ws.cell(row=4, column=1, value="Version Mode")
    version_mode = "Latest Only" if latest_only else "All Versions"
    summary_ws.cell(row=4, column=2, value=version_mode)

    auto_adjust_column_widths(summary_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_snapshot_report_excel(
    data: list[dict[str, Any]],
) -> BytesIO:
    """Create an Excel file for SNAPSHOT dependency report.

    Args:
        data: List of SNAPSHOT dependency records

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SNAPSHOT Dependencies"

    # Headers
    headers = ["Application", "App Version", "Dependency", "Dependency Version"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    # Data
    for idx, record in enumerate(data, start=2):
        ws.cell(row=idx, column=1, value=record["application"])
        ws.cell(row=idx, column=2, value=record["app_version"])
        ws.cell(row=idx, column=3, value=record["dependency"])
        ws.cell(row=idx, column=4, value=record["dep_version"])

    auto_adjust_column_widths(ws)

    # Summary
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Total SNAPSHOT Dependencies")
    summary_ws.cell(row=1, column=2, value=len(data))

    unique_apps = len({r["application"] for r in data})
    summary_ws.cell(row=2, column=1, value="Affected Applications")
    summary_ws.cell(row=2, column=2, value=unique_apps)

    unique_deps = len({r["dependency"] for r in data})
    summary_ws.cell(row=3, column=1, value="Unique SNAPSHOT Dependencies")
    summary_ws.cell(row=3, column=2, value=unique_deps)

    auto_adjust_column_widths(summary_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_self_dependency_report_excel(
    data: list[dict[str, Any]],
) -> BytesIO:
    """Create an Excel file for self-dependency report.

    Args:
        data: List of self-dependency records

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Self Dependencies"

    # Headers
    headers = ["Project Name", "Version", "Relationship Type"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    # Data
    for idx, record in enumerate(data, start=2):
        ws.cell(row=idx, column=1, value=record["project_name"])
        ws.cell(row=idx, column=2, value=record["version"])
        ws.cell(row=idx, column=3, value=record["relationship_type"])

    auto_adjust_column_widths(ws)

    # Summary
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Total Self Dependencies")
    summary_ws.cell(row=1, column=2, value=len(data))

    unique_projects = len({r["project_name"] for r in data})
    summary_ws.cell(row=2, column=1, value="Affected Projects")
    summary_ws.cell(row=2, column=2, value=unique_projects)

    auto_adjust_column_widths(summary_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_multi_version_dependency_report_excel(
    data: dict[str, Any],
) -> BytesIO:
    """Create an Excel file for multi-version dependency source report.

    This report shows dependencies with multiple versions and the applications
    that contributed each version to the target project's dependency graph.

    Args:
        data: Dict with 'target' info and 'multi_version_dependencies' list

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Version Dependencies"

    # Headers
    headers = [
        "Dependency Project",
        "Dependency Version",
        "Contributing Application",
        "Application Version",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    # Data - flatten the nested structure
    row_idx = 2
    multi_deps = data.get("multi_version_dependencies", [])

    for dep in multi_deps:
        dep_project = dep["dependency_project"]
        for version_info in dep["versions"]:
            dep_version = version_info["version"]
            apps = version_info["contributing_applications"]

            if apps:
                for app in apps:
                    ws.cell(row=row_idx, column=1, value=dep_project)
                    ws.cell(row=row_idx, column=2, value=dep_version)
                    ws.cell(row=row_idx, column=3, value=app["project_name"])
                    ws.cell(row=row_idx, column=4, value=app["version"])
                    row_idx += 1
            else:
                # No contributing apps found (scan_ids may not match)
                ws.cell(row=row_idx, column=1, value=dep_project)
                ws.cell(row=row_idx, column=2, value=dep_version)
                ws.cell(row=row_idx, column=3, value="(unknown source)")
                ws.cell(row=row_idx, column=4, value="-")
                row_idx += 1

    auto_adjust_column_widths(ws)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    target = data.get("target", {})

    summary_ws.cell(row=1, column=1, value="Target Project")
    summary_ws.cell(row=1, column=2, value=target.get("project_name", "N/A"))

    summary_ws.cell(row=2, column=1, value="Target Version")
    summary_ws.cell(row=2, column=2, value=target.get("version", "N/A"))

    summary_ws.cell(row=3, column=1, value="Scan IDs Contributing")
    summary_ws.cell(row=3, column=2, value=target.get("scan_ids_count", 0))

    summary_ws.cell(row=4, column=1, value="Dependencies with Multiple Versions")
    summary_ws.cell(row=4, column=2, value=len(multi_deps))

    # Count total versions across all multi-version deps
    total_versions = sum(dep["version_count"] for dep in multi_deps)
    summary_ws.cell(row=5, column=1, value="Total Conflicting Versions")
    summary_ws.cell(row=5, column=2, value=total_versions)

    # Count unique contributing applications
    all_apps = set()
    for dep in multi_deps:
        for version_info in dep["versions"]:
            for app in version_info["contributing_applications"]:
                all_apps.add((app["project_name"], app["version"]))
    summary_ws.cell(row=6, column=1, value="Unique Contributing Applications")
    summary_ws.cell(row=6, column=2, value=len(all_apps))

    auto_adjust_column_widths(summary_ws)

    # Detailed breakdown sheet - one row per dependency with version counts
    detail_ws = wb.create_sheet("Dependency Summary")

    detail_headers = ["Dependency Project", "Version Count", "Versions"]
    for col, header in enumerate(detail_headers, start=1):
        detail_ws.cell(row=1, column=col, value=header)
    style_header_row(detail_ws, len(detail_headers))

    for idx, dep in enumerate(multi_deps, start=2):
        detail_ws.cell(row=idx, column=1, value=dep["dependency_project"])
        detail_ws.cell(row=idx, column=2, value=dep["version_count"])
        versions_str = ", ".join(v["version"] for v in dep["versions"])
        detail_ws.cell(row=idx, column=3, value=versions_str)

    auto_adjust_column_widths(detail_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_multi_version_deps_excel(
    data: dict[str, Any],
) -> BytesIO:
    """Create an Excel file for library version usage report.

    This report shows all versions of a library and which projects use each.
    Use case: Understanding library adoption patterns across the organization.

    Args:
        data: Dict with 'library' info, 'total_dependants', and 'versions' list

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Version Usage"

    # Headers
    headers = [
        "Library Version",
        "Dependant Count",
        "Dependant Project",
        "Dependant Version",
        "Project Group",
        "Is Internal",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    # Data - flatten the nested structure
    row_idx = 2
    versions = data.get("versions", [])

    for ver_info in versions:
        version = ver_info.get("version", "")
        dependant_count = ver_info.get("dependant_count", 0)
        dependants = ver_info.get("dependants", [])

        if dependants:
            for dep in dependants:
                ws.cell(row=row_idx, column=1, value=version)
                ws.cell(row=row_idx, column=2, value=dependant_count)
                ws.cell(row=row_idx, column=3, value=dep.get("project_name", ""))
                ws.cell(row=row_idx, column=4, value=dep.get("version", ""))
                ws.cell(row=row_idx, column=5, value=dep.get("project_group", ""))
                ws.cell(row=row_idx, column=6, value="Yes" if dep.get("is_internal") else "No")
                row_idx += 1
        else:
            # No dependants for this version
            ws.cell(row=row_idx, column=1, value=version)
            ws.cell(row=row_idx, column=2, value=0)
            ws.cell(row=row_idx, column=3, value="(no direct dependants)")
            ws.cell(row=row_idx, column=4, value="-")
            ws.cell(row=row_idx, column=5, value="-")
            ws.cell(row=row_idx, column=6, value="-")
            row_idx += 1

    auto_adjust_column_widths(ws)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    library_info = data.get("library", {})

    summary_ws.cell(row=1, column=1, value="Library")
    summary_ws.cell(row=1, column=2, value=library_info.get("project_name", "N/A"))

    summary_ws.cell(row=2, column=1, value="Total Versions")
    summary_ws.cell(row=2, column=2, value=library_info.get("total_versions", 0))

    summary_ws.cell(row=3, column=1, value="Total Dependants")
    summary_ws.cell(row=3, column=2, value=data.get("total_dependants", 0))

    auto_adjust_column_widths(summary_ws)

    # Version breakdown sheet - one row per version with counts
    detail_ws = wb.create_sheet("Version Summary")

    detail_headers = ["Version", "Dependant Count", "Is Internal"]
    for col, header in enumerate(detail_headers, start=1):
        detail_ws.cell(row=1, column=col, value=header)
    style_header_row(detail_ws, len(detail_headers))

    for idx, ver_info in enumerate(versions, start=2):
        detail_ws.cell(row=idx, column=1, value=ver_info.get("version", ""))
        detail_ws.cell(row=idx, column=2, value=ver_info.get("dependant_count", 0))
        detail_ws.cell(row=idx, column=3, value="Yes" if ver_info.get("is_internal") else "No")

    auto_adjust_column_widths(detail_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_non_semver_report_excel(
    data: list[dict[str, Any]],
) -> BytesIO:
    """Create an Excel file for non-SemVer versions report.

    Args:
        data: List of non-SemVer version records with project_name, version, reason

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Non-SemVer Versions"

    # Headers
    headers = ["Project Name", "Version", "Reason", "Labels"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    # Data
    for idx, record in enumerate(data, start=2):
        ws.cell(row=idx, column=1, value=record["project_name"])
        ws.cell(row=idx, column=2, value=record["version"])
        ws.cell(row=idx, column=3, value=record["reason"])
        labels = record.get("labels", [])
        ws.cell(row=idx, column=4, value=", ".join(labels) if labels else "")

    auto_adjust_column_widths(ws)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Total Non-SemVer Versions")
    summary_ws.cell(row=1, column=2, value=len(data))

    unique_projects = len({r["project_name"] for r in data})
    summary_ws.cell(row=2, column=1, value="Affected Projects")
    summary_ws.cell(row=2, column=2, value=unique_projects)

    # Count by reason
    reason_counts: dict[str, int] = {}
    for record in data:
        reason = record["reason"]
        reason_counts[reason] = reason_counts.get(reason, 0) + 1

    summary_ws.cell(row=4, column=1, value="Breakdown by Reason")
    style_header_row(summary_ws, 2)

    row_idx = 5
    for reason, count in sorted(reason_counts.items(), key=lambda x: -x[1]):
        summary_ws.cell(row=row_idx, column=1, value=reason)
        summary_ws.cell(row=row_idx, column=2, value=count)
        row_idx += 1

    auto_adjust_column_widths(summary_ws)

    # By-reason sheet for detailed breakdown
    by_reason_ws = wb.create_sheet("By Reason")
    reason_headers = ["Reason", "Project Name", "Version"]
    for col, header in enumerate(reason_headers, start=1):
        by_reason_ws.cell(row=1, column=col, value=header)
    style_header_row(by_reason_ws, len(reason_headers))

    # Sort data by reason for grouping
    sorted_data = sorted(data, key=lambda x: (x["reason"], x["project_name"], x["version"]))
    for idx, record in enumerate(sorted_data, start=2):
        by_reason_ws.cell(row=idx, column=1, value=record["reason"])
        by_reason_ws.cell(row=idx, column=2, value=record["project_name"])
        by_reason_ws.cell(row=idx, column=3, value=record["version"])

    auto_adjust_column_widths(by_reason_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_dependants_report_excel(
    data: dict[str, Any],
    longest_only: bool = True,
) -> BytesIO:
    """Create an Excel file for dependants report with partitions and paths.

    Args:
        data: Dict with 'target', 'stats', and 'dependants' list
        longest_only: If True, report contains only longest paths

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dependants"

    # Headers
    headers = [
        "Dependant Project",
        "Version",
        "Partition (Longest Path)",
        "Max Path Edges",
        "Path Count",
        "Labels",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    # Data
    dependants = data.get("dependants", [])
    for idx, dep in enumerate(dependants, start=2):
        ws.cell(row=idx, column=1, value=dep.get("project_name", ""))
        ws.cell(row=idx, column=2, value=dep.get("version", ""))
        ws.cell(row=idx, column=3, value=dep.get("partition", 0))
        ws.cell(row=idx, column=4, value=dep.get("max_path_edges", 0))
        ws.cell(row=idx, column=5, value=dep.get("path_count", 0))
        ws.cell(row=idx, column=6, value=", ".join(dep.get("labels", [])))

    auto_adjust_column_widths(ws)

    # Paths sheet - detailed path information
    paths_ws = wb.create_sheet("Dependency Paths")
    path_headers = ["Dependant", "Version", "Path #", "Path Length", "Path"]
    for col, header in enumerate(path_headers, start=1):
        paths_ws.cell(row=1, column=col, value=header)
    style_header_row(paths_ws, len(path_headers))

    row_idx = 2
    for dep in dependants:
        for path_num, path in enumerate(dep.get("paths", []), start=1):
            paths_ws.cell(row=row_idx, column=1, value=dep.get("project_name", ""))
            paths_ws.cell(row=row_idx, column=2, value=dep.get("version", ""))
            paths_ws.cell(row=row_idx, column=3, value=path_num)
            paths_ws.cell(row=row_idx, column=4, value=len(path))
            paths_ws.cell(row=row_idx, column=5, value=" -> ".join(path))
            row_idx += 1

    auto_adjust_column_widths(paths_ws)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    target = data.get("target", {})
    stats = data.get("stats", {})

    summary_ws.cell(row=1, column=1, value="Target Project")
    summary_ws.cell(row=1, column=2, value=target.get("project_name", "N/A"))

    summary_ws.cell(row=2, column=1, value="Target Version")
    summary_ws.cell(row=2, column=2, value=target.get("version", "N/A"))

    summary_ws.cell(row=3, column=1, value="Total Dependants")
    summary_ws.cell(row=3, column=2, value=stats.get("total_dependants", 0))

    summary_ws.cell(row=4, column=1, value="Max Partition (Longest Path)")
    summary_ws.cell(row=4, column=2, value=stats.get("max_partition", 0))

    summary_ws.cell(row=5, column=1, value="Unique Projects")
    summary_ws.cell(row=5, column=2, value=stats.get("unique_projects", 0))

    summary_ws.cell(row=6, column=1, value="Path Mode")
    summary_ws.cell(row=6, column=2, value="Longest only" if longest_only else "All paths")

    auto_adjust_column_widths(summary_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_vulnerabilities_excel(
    vulnerabilities: list[dict[str, Any]],
    internal_only: bool = False,
) -> BytesIO:
    """Create an Excel file for vulnerabilities report.

    Args:
        vulnerabilities: List of vulnerability dicts
        internal_only: If True, indicates internal-only filter was applied

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    # Headers
    headers = [
        "ID",
        "Severity",
        "CVSS Score",
        "Title",
        "CWE",
        "Published Date",
        "Affected Project",
        "Affected Version",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    # Data - one row per affected version
    row_idx = 2
    for vuln in vulnerabilities:
        affected_versions = vuln.get("affected_versions", [])
        if not affected_versions:
            # Still show vulnerability even without affected versions
            ws.cell(row=row_idx, column=1, value=vuln.get("defect_id", ""))
            ws.cell(row=row_idx, column=2, value=vuln.get("severity", ""))
            ws.cell(row=row_idx, column=3, value=vuln.get("cvss_score", 0))
            ws.cell(row=row_idx, column=4, value=vuln.get("title", ""))
            ws.cell(row=row_idx, column=5, value=vuln.get("cwe_id", ""))
            ws.cell(row=row_idx, column=6, value=vuln.get("published_date", ""))
            ws.cell(row=row_idx, column=7, value="")
            ws.cell(row=row_idx, column=8, value="")
            row_idx += 1
        else:
            for affected in affected_versions:
                ws.cell(row=row_idx, column=1, value=vuln.get("defect_id", ""))
                ws.cell(row=row_idx, column=2, value=vuln.get("severity", ""))
                ws.cell(row=row_idx, column=3, value=vuln.get("cvss_score", 0))
                ws.cell(row=row_idx, column=4, value=vuln.get("title", ""))
                ws.cell(row=row_idx, column=5, value=vuln.get("cwe_id", ""))
                ws.cell(row=row_idx, column=6, value=vuln.get("published_date", ""))
                ws.cell(row=row_idx, column=7, value=affected.get("project_name", ""))
                ws.cell(row=row_idx, column=8, value=affected.get("version", ""))
                row_idx += 1

    auto_adjust_column_widths(ws)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Total Vulnerabilities")
    summary_ws.cell(row=1, column=2, value=len(vulnerabilities))

    # Count by severity
    severity_counts: dict[str, int] = {}
    total_affected = 0
    for vuln in vulnerabilities:
        sev = vuln.get("severity", "UNKNOWN")
        severity_counts[sev] = severity_counts.get(sev, 0) + 1
        total_affected += len(vuln.get("affected_versions", []))

    summary_ws.cell(row=2, column=1, value="Total Affected Versions")
    summary_ws.cell(row=2, column=2, value=total_affected)

    summary_ws.cell(row=3, column=1, value="Filter Mode")
    summary_ws.cell(row=3, column=2, value="Internal Only" if internal_only else "All")

    row_idx = 5
    summary_ws.cell(row=row_idx, column=1, value="By Severity")
    row_idx += 1
    for severity in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        count = severity_counts.get(severity, 0)
        if count > 0:
            summary_ws.cell(row=row_idx, column=1, value=severity)
            summary_ws.cell(row=row_idx, column=2, value=count)
            row_idx += 1

    auto_adjust_column_widths(summary_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_vulnerability_dependants_excel(
    vulnerability: dict[str, Any],
    dependants: list[dict[str, Any]],
    internal_only: bool = False,
) -> BytesIO:
    """Create an Excel file for vulnerability dependants report.

    Args:
        vulnerability: The vulnerability details
        dependants: List of dependant dicts with partition info
        internal_only: If True, indicates internal-only filter was applied

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Affected Dependants"

    # Headers
    headers = [
        "Partition",
        "Project Name",
        "Version",
        "Is Internal",
        "Affected Via (Project)",
        "Affected Via (Version)",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    # Data - one row per dependant
    row_idx = 2
    for dep in dependants:
        affected_by = dep.get("affected_by", [])
        affected_via_projects = ", ".join(a.get("project_name", "") for a in affected_by)
        affected_via_versions = ", ".join(a.get("version", "") for a in affected_by)

        ws.cell(row=row_idx, column=1, value=dep.get("partition", 0))
        ws.cell(row=row_idx, column=2, value=dep.get("project_name", ""))
        ws.cell(row=row_idx, column=3, value=dep.get("version", ""))
        ws.cell(row=row_idx, column=4, value="Yes" if dep.get("is_internal") else "No")
        ws.cell(row=row_idx, column=5, value=affected_via_projects)
        ws.cell(row=row_idx, column=6, value=affected_via_versions)
        row_idx += 1

    auto_adjust_column_widths(ws)

    # Vulnerability details sheet
    vuln_ws = wb.create_sheet("Vulnerability")
    vuln_ws.cell(row=1, column=1, value="ID")
    vuln_ws.cell(row=1, column=2, value=vulnerability.get("defect_id", ""))

    vuln_ws.cell(row=2, column=1, value="Severity")
    vuln_ws.cell(row=2, column=2, value=vulnerability.get("severity", ""))

    vuln_ws.cell(row=3, column=1, value="CVSS Score")
    vuln_ws.cell(row=3, column=2, value=vulnerability.get("cvss_score", 0))

    vuln_ws.cell(row=4, column=1, value="Title")
    vuln_ws.cell(row=4, column=2, value=vulnerability.get("title", ""))

    vuln_ws.cell(row=5, column=1, value="CWE")
    vuln_ws.cell(row=5, column=2, value=vulnerability.get("cwe_id", ""))

    vuln_ws.cell(row=6, column=1, value="Published Date")
    vuln_ws.cell(row=6, column=2, value=vulnerability.get("published_date", ""))

    vuln_ws.cell(row=7, column=1, value="Description")
    vuln_ws.cell(row=7, column=2, value=vulnerability.get("description", ""))

    auto_adjust_column_widths(vuln_ws)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.cell(row=1, column=1, value="Total Dependants")
    summary_ws.cell(row=1, column=2, value=len(dependants))

    max_partition = max((d.get("partition", 0) for d in dependants), default=0)
    summary_ws.cell(row=2, column=1, value="Max Partition")
    summary_ws.cell(row=2, column=2, value=max_partition)

    unique_projects = len({d["project_name"] for d in dependants})
    summary_ws.cell(row=3, column=1, value="Unique Projects")
    summary_ws.cell(row=3, column=2, value=unique_projects)

    summary_ws.cell(row=4, column=1, value="Filter Mode")
    summary_ws.cell(row=4, column=2, value="Internal Only" if internal_only else "All")

    # Partition breakdown
    partition_counts: dict[int, int] = {}
    for dep in dependants:
        p = dep.get("partition", 0)
        partition_counts[p] = partition_counts.get(p, 0) + 1

    row_idx = 6
    summary_ws.cell(row=row_idx, column=1, value="By Partition")
    row_idx += 1
    for partition in sorted(partition_counts.keys()):
        summary_ws.cell(row=row_idx, column=1, value=f"Partition {partition}")
        summary_ws.cell(row=row_idx, column=2, value=partition_counts[partition])
        row_idx += 1

    auto_adjust_column_widths(summary_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_centrality_excel(
    centrality_data: list[dict[str, Any]],
) -> BytesIO:
    """Create Excel file for centrality report.

    Args:
        centrality_data: List of dicts with inDegree, outDegree, project info

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Centrality"

    # Headers
    headers = ["inDegree", "outDegree", "Project Name", "Project Group", "Version"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Data
    row_idx = 2
    for item in centrality_data:
        ws.cell(row=row_idx, column=1, value=item.get("inDegree", 0))
        ws.cell(row=row_idx, column=2, value=item.get("outDegree", 0))
        ws.cell(row=row_idx, column=3, value=item.get("project_name", ""))
        ws.cell(row=row_idx, column=4, value=item.get("project_group", ""))
        ws.cell(row=row_idx, column=5, value=item.get("version_name", ""))
        row_idx += 1

    auto_adjust_column_widths(ws)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")

    total_libs = len(centrality_data)
    total_in = sum(d.get("inDegree", 0) for d in centrality_data)
    total_out = sum(d.get("outDegree", 0) for d in centrality_data)
    max_in = max((d.get("inDegree", 0) for d in centrality_data), default=0)
    max_out = max((d.get("outDegree", 0) for d in centrality_data), default=0)

    summary_ws.cell(row=1, column=1, value="Total Internal Libraries")
    summary_ws.cell(row=1, column=2, value=total_libs)

    summary_ws.cell(row=2, column=1, value="Total Inward Connections")
    summary_ws.cell(row=2, column=2, value=total_in)

    summary_ws.cell(row=3, column=1, value="Total Outward Connections")
    summary_ws.cell(row=3, column=2, value=total_out)

    summary_ws.cell(row=4, column=1, value="Max inDegree")
    summary_ws.cell(row=4, column=2, value=max_in)

    summary_ws.cell(row=5, column=1, value="Max outDegree")
    summary_ws.cell(row=5, column=2, value=max_out)

    # Top 10 by inDegree
    summary_ws.cell(row=7, column=1, value="Top 10 by inDegree")
    summary_ws.cell(row=7, column=1).font = Font(bold=True)

    sorted_by_in = sorted(centrality_data, key=lambda x: x.get("inDegree", 0), reverse=True)[:10]
    for i, item in enumerate(sorted_by_in, 8):
        summary_ws.cell(row=i, column=1, value=f"{item['project_name']}:{item['version_name']}")
        summary_ws.cell(row=i, column=2, value=item.get("inDegree", 0))

    # Top 10 by outDegree
    row_offset = 8 + len(sorted_by_in) + 2
    summary_ws.cell(row=row_offset, column=1, value="Top 10 by outDegree")
    summary_ws.cell(row=row_offset, column=1).font = Font(bold=True)

    sorted_by_out = sorted(centrality_data, key=lambda x: x.get("outDegree", 0), reverse=True)[:10]
    for i, item in enumerate(sorted_by_out, row_offset + 1):
        summary_ws.cell(row=i, column=1, value=f"{item['project_name']}:{item['version_name']}")
        summary_ws.cell(row=i, column=2, value=item.get("outDegree", 0))

    auto_adjust_column_widths(summary_ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
